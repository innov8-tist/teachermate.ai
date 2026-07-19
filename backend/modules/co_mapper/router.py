"""
CO Mapper Router
API endpoints for CO mapping functionality
"""

from fastapi import APIRouter, Depends, File, UploadFile, Form, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from pathlib import Path
from io import BytesIO
import traceback
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from db_service.db import get_db
from modules.auth.dependencies import get_current_teacher
from modules.auth.models import Teacher

from .service import CoMapperService
from .schemas import *


router = APIRouter(prefix="", tags=["co_mapper"])

TEMP_FOLDER = Path(tempfile.gettempdir()) / "co_images"
TEMP_FOLDER.mkdir(parents=True, exist_ok=True)


def get_co_service(db: Session = Depends(get_db)) -> CoMapperService:
    """Dependency to get CO mapper service"""
    return CoMapperService(db)


@router.get("/subject_fetch/{semester}")
def fetch_subjects_by_semester(
    semester: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get all subjects for a given semester"""
    subjects = service.get_subjects_by_semester(semester)
    return subjects


@router.post("/co_creation", response_model=CoCreationResponse)
async def create_co_template(
    subject_name: str = Form(...),
    sem: int = Form(...),
    ia_number: int = Form(...),
    student_count: int = Form(...),
    co_image: UploadFile = File(...),
    current_teacher: Teacher = Depends(get_current_teacher),
    service: CoMapperService = Depends(get_co_service)
):
    """Create a new CO mapping template from uploaded question paper image"""
    print(f"Received Form Data:")
    print(f"subject_name: {subject_name} (type: {type(subject_name)})")
    print(f"sem: {sem} (type: {type(sem)})")
    print(f"ia_number: {ia_number} (type: {type(ia_number)})")
    print(f"student_count: {student_count} (type: {type(student_count)})")
    
    file_content = await co_image.read()
    
    try:
        created_subject = service.create_co_template(
            subject_name=subject_name,
            sem=sem,
            ia_number=ia_number,
            student_count=student_count,
            teacher_id=current_teacher.id,
            image_content=file_content,
            image_filename=co_image.filename,
            temp_folder=TEMP_FOLDER
        )
        
        return CoCreationResponse(
            status="success",
            message="CO created successfully",
            data={
                "id": created_subject.id,
                "subject_name": created_subject.name,
                "branch": created_subject.branch,
                "semester": created_subject.sem,
                "ia": created_subject.ia,
                "image_path": created_subject.image_path
            }
        )
    except ValueError as e:
        return CoCreationResponse(
            status="error",
            message=str(e)
        )
    except Exception as e:
        return CoCreationResponse(
            status="error",
            message=f"Failed to create CO: {str(e)}"
        )


@router.get("/co_fetch/{teacher_id}")
def get_teacher_co_templates(
    teacher_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get all CO templates created by a teacher"""
    all_co = service.get_templates_by_teacher(teacher_id)
    return all_co


@router.get('/co_questions/{subject_id}', response_model=AllQuestionsResponse)
def get_co_questions(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get all question numbers for a CO template"""
    questions = service.get_questions_list(subject_id)
    return AllQuestionsResponse(all_questions=questions)


@router.get("/co_fetch_details/{subject_id}")
def get_co_details(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get CO question mappings for a template"""
    details = service.get_co_details(subject_id)
    return details


@router.get("/co_subject_info/{subject_id}")
def get_co_subject_info(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get subject/template information"""
    subject_info = service.get_template_info(subject_id)
    return subject_info


@router.get("/co_download_excel/{subject_id}")
def download_co_excel(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """
    Generate and download Excel file with student marks mapped to COs
    Format: Register Number | CO1 (Q1, Q2, Total) | CO2 (Q3, Q4, Total) | ...
    """
    try:
        subject_info = service.get_template_info(subject_id)
        if not subject_info:
            return {"status": "error", "message": "Subject not found"}
        
        data = service.get_excel_data(subject_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CO Mapping"

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate total columns needed
        total_cols = 2  # Registration Number + Name columns
        for co, questions in data['co_structure'].items():
            total_cols += len(questions) + 1  # questions + total column
        
        # ROW 1: Title row (merged across all columns)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{subject_info['name']} - {subject_info['ia']} - CO Mapping"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        
        # ROW 2 & 3: Headers
        # Column A: Registration Number (merged rows 2-3)
        cell = ws.cell(row=2, column=1)
        cell.value = "Register Number"
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        
        # Column B: Student Name (merged rows 2-3)
        cell = ws.cell(row=2, column=2)
        cell.value = "Student Name"
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        
        # CO Headers (row 2) and Question numbers (row 3)
        col_idx = 3  # Start after Name column
        
        for co, questions in data['co_structure'].items():
            start_col = col_idx
            end_col = col_idx + len(questions)  # Include total column
            
            # Set CO name BEFORE merging
            cell = ws.cell(row=2, column=start_col)
            cell.value = co
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            
            # Add borders to merged cells in row 2
            for c in range(start_col, end_col + 1):
                ws.cell(row=2, column=c).border = border
            
            # Question numbers in row 3
            for question in questions:
                cell = ws.cell(row=3, column=col_idx)
                cell.value = question
                cell.font = Font(bold=True, size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                # Set column width
                col_letter = ws.cell(row=3, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 8
                col_idx += 1
            
            # Total column for this CO
            cell = ws.cell(row=3, column=col_idx)
            cell.value = f"Total {co}"
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            col_letter = ws.cell(row=3, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 10
            col_idx += 1
        
        # Data rows with student information
        for row_idx, student in enumerate(data['students'], start=4):
            col_idx = 1
            
            # Registration number
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = student['regno']
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col_idx += 1
            
            # Student name
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = student.get('name', '')
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            col_idx += 1
            
            # Marks for each CO
            for co, questions in data['co_structure'].items():
                student_co_marks = student['marks'].get(co, {})
                
                # Individual question marks
                for question in questions:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = student_co_marks.get(question, 0)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    col_idx += 1
                
                # Total for this CO
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = student_co_marks.get('total', 0)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
                col_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 18  # Registration Number
        ws.column_dimensions['B'].width = 25  # Student Name
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"{subject_info['name']}_{subject_info['ia']}_CO_Mapping.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error generating Excel: {str(e)}")
        traceback.print_exc()
        return {"status": "error", "message": f"Failed to generate Excel: {str(e)}"}


@router.delete("/co_delete/{subject_id}", response_model=DeleteResponse)
def delete_co_template(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Delete a CO template and all associated data (cascading delete)"""
    success = service.delete_co_template(subject_id)
    if success:
        return DeleteResponse(status="success", message="CO deleted successfully")
    else:
        return DeleteResponse(status="error", message="CO not found")


@router.post("/student_sheet_upload", response_model=StudentSheetUploadResponse)
async def upload_student_sheet(
    subject_id: int = Form(...),
    student_image: UploadFile = File(...),
    current_teacher: Teacher = Depends(get_current_teacher),
    service: CoMapperService = Depends(get_co_service)
):
    """
    Upload and process student answer sheet
    Extracts registration number and marks, saves to CO Mapper
    Performs reverse sync to Evaluation system if schema exists
    """
    import time
    import traceback
    
    start_time = time.time()
    
    try:
        print(f"\n{'='*80}")
        print(f"🚀 STUDENT SHEET UPLOAD STARTED")
        print(f"📊 Subject ID: {subject_id}")
        print(f"📁 Filename: {student_image.filename}")
        print(f"📏 Content Type: {student_image.content_type}")
        print(f"⏰ Start Time: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*80}")
        
        # Step 1: Read file content
        read_start = time.time()
        print(f"📖 [STEP 1] Reading file content...")
        file_content = await student_image.read()
        read_time = time.time() - read_start
        
        print(f"✅ File read completed in {read_time:.2f}s")
        print(f"📏 File size: {len(file_content):,} bytes ({len(file_content)/(1024*1024):.2f} MB)")
        
        # File size validation
        max_size = 50 * 1024 * 1024  # 50MB
        if len(file_content) > max_size:
            error_msg = f"File too large: {len(file_content):,} bytes (max: {max_size:,} bytes)"
            print(f"❌ {error_msg}")
            return StudentSheetUploadResponse(
                status="error",
                message=error_msg,
                evaluation_synced=False
            )
        
        # Step 2: Process student sheet
        process_start = time.time()
        print(f"🔄 [STEP 2] Processing student sheet...")
        
        result = service.process_student_sheet(
            template_id=subject_id,
            image_content=file_content,
            image_filename=student_image.filename,
            temp_folder=TEMP_FOLDER
        )
        
        process_time = time.time() - process_start
        total_time = time.time() - start_time
        
        print(f"✅ Processing completed in {process_time:.2f}s")
        print(f"📊 Extracted data: {result.get('data', {})}")
        print(f"🔄 Evaluation synced: {result['evaluation_synced']}")
        print(f"\n{'='*80}")
        print(f"🎉 UPLOAD SUCCESS - Total time: {total_time:.2f}s")
        print(f"{'='*80}\n")
        
        return StudentSheetUploadResponse(
            status="success",
            message="Student answer sheet uploaded, processed, extracted, and saved to database successfully",
            evaluation_synced=result['evaluation_synced'],
            data=result['data']
        )
        
    except Exception as e:
        error_time = time.time() - start_time
        
        print(f"\n{'='*80}")
        print(f"❌ UPLOAD FAILED after {error_time:.2f}s")
        print(f"🚨 Error Type: {type(e).__name__}")
        print(f"💬 Error Message: {str(e)}")
        print(f"{'='*80}")
        print("📋 Full Traceback:")
        traceback.print_exc()
        print(f"{'='*80}\n")
        
        return StudentSheetUploadResponse(
            status="error",
            message=f"Failed to upload: {str(e)}",
            evaluation_synced=False
        )


@router.get("/students_by_subject/{subject_id}")
def get_students_by_subject(
    subject_id: int,
    service: CoMapperService = Depends(get_co_service)
):
    """Get all students who have submitted answers for a CO template"""
    students = service.get_students_by_template(subject_id)
    return students


@router.get("/student_marks/{subject_id}/{regno}")
def get_student_marks_detail(
    subject_id: int,
    regno: str,
    service: CoMapperService = Depends(get_co_service)
):
    """Get detailed marks for a specific student"""
    marks = service.get_student_marks(subject_id, regno)
    return marks


@router.delete("/student_marks/{subject_id}/{regno}", response_model=DeleteResponse)
def delete_student_marks(
    subject_id: int,
    regno: str,
    service: CoMapperService = Depends(get_co_service)
):
    """
    Delete student marks from CO Mapper
    Also removes corresponding evaluation records
    """
    success = service.delete_student_marks(subject_id, regno)
    if success:
        return DeleteResponse(
            status="success",
            message="Student marks and evaluation records deleted successfully"
        )
    else:
        return DeleteResponse(
            status="error",
            message="Student marks not found"
        )
